from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, current_app, jsonify
from flask_login import login_required, current_user
from models import Athlete, AthleteGroup, Group, Payment, AthleteDocument
from app import db
from datetime import date, datetime
from routes.decorators import permission_required, club_required, current_club_id, scoped_or_404
import os, uuid, io

athletes_bp = Blueprint('athletes', __name__)

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ── Lista (admin, entrenador, recepcion pueden ver) ──────────────────────────
@athletes_bp.route('/athletes')
@login_required
@club_required
@permission_required('athletes.view')
def index():
    search   = request.args.get('search', '')
    status   = request.args.get('status', '')
    group_id = request.args.get('group', '')
    query = Athlete.query.filter_by(club_id=current_club_id())
    if search:
        query = query.filter(
            (Athlete.first_name.ilike(f'%{search}%')) |
            (Athlete.last_name.ilike(f'%{search}%'))  |
            (Athlete.document_number.ilike(f'%{search}%'))
        )
    if status:
        query = query.filter_by(status=status)
    if group_id:
        query = query.join(AthleteGroup).filter(
            AthleteGroup.group_id == int(group_id), AthleteGroup.active == True
        )
    athletes = query.order_by(Athlete.last_name).all()
    groups   = Group.query.filter_by(active=True, club_id=current_club_id()).all()
    return render_template('athletes/index.html', athletes=athletes, groups=groups,
                           search=search, status_filter=status, group_filter=group_id)


# ── Detalle (admin, entrenador, recepcion solo lectura) ──────────────────────
@athletes_bp.route('/athletes/<int:id>')
@login_required
@club_required
@permission_required('athletes.view')
def detail(id):
    athlete = scoped_or_404(Athlete, id)
    groups  = Group.query.filter_by(active=True, club_id=current_club_id()).all()
    payments  = sorted(athlete.payments, key=lambda p: (p.date or date.min), reverse=True)
    documents = sorted(athlete.documents, key=lambda d: d.uploaded_at or datetime.min, reverse=True)
    return render_template('athletes/detail.html', athlete=athlete, groups=groups,
                           payments=payments, documents=documents)


# ── Crear / Editar / Eliminar (NO recepcion) ─────────────────────────────────
@athletes_bp.route('/athletes/new', methods=['GET', 'POST'])
@login_required
@club_required
@permission_required('athletes.manage')
def new():
    groups = Group.query.filter_by(active=True, club_id=current_club_id()).all()
    if request.method == 'POST':
        athlete = Athlete(
            club_id=current_club_id(),
            first_name=request.form.get('first_name'),
            last_name=request.form.get('last_name'),
            document_type=request.form.get('document_type', 'CC'),
            document_number=request.form.get('document_number'),
            birth_date=_parse_date(request.form.get('birth_date')),
            gender=request.form.get('gender'),
            phone=request.form.get('phone'),
            email=request.form.get('email'),
            address=request.form.get('address'),
            city=request.form.get('city'),
            emergency_contact=request.form.get('emergency_contact'),
            emergency_phone=request.form.get('emergency_phone'),
            status=request.form.get('status', 'activo'),
            position=request.form.get('position') or None,
            entry_date=_parse_date(request.form.get('entry_date')) or date.today(),
            notes=request.form.get('notes'),
        )
        db.session.add(athlete)
        db.session.flush()
        for gid in request.form.getlist('groups'):
            db.session.add(AthleteGroup(athlete_id=athlete.id, group_id=int(gid)))
        # Payments
        concepts = request.form.getlist('pay_concept')
        amounts  = request.form.getlist('pay_amount')
        ptypes   = request.form.getlist('pay_type')
        pdates   = request.form.getlist('pay_date')
        for i in range(len(concepts)):
            if concepts[i]:
                db.session.add(Payment(
                    athlete_id=athlete.id,
                    concept=concepts[i],
                    amount=float(amounts[i] or 0),
                    type=ptypes[i] if i < len(ptypes) else 'cargo',
                    date=_parse_date(pdates[i]) if i < len(pdates) else date.today(),
                    status='pendiente',
                ))
        db.session.commit()
        flash('Deportista creado exitosamente.', 'success')
        return redirect(url_for('athletes.detail', id=athlete.id))
    return render_template('athletes/form.html', athlete=None, groups=groups)


@athletes_bp.route('/athletes/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@club_required
@permission_required('athletes.manage')
def edit(id):
    athlete = scoped_or_404(Athlete, id)
    groups  = Group.query.filter_by(active=True, club_id=current_club_id()).all()
    if request.method == 'POST':
        athlete.first_name        = request.form.get('first_name')
        athlete.last_name         = request.form.get('last_name')
        athlete.document_type     = request.form.get('document_type', 'CC')
        athlete.document_number   = request.form.get('document_number')
        athlete.birth_date        = _parse_date(request.form.get('birth_date'))
        athlete.gender            = request.form.get('gender')
        athlete.phone             = request.form.get('phone')
        athlete.email             = request.form.get('email')
        athlete.address           = request.form.get('address')
        athlete.city              = request.form.get('city')
        athlete.emergency_contact = request.form.get('emergency_contact')
        athlete.emergency_phone   = request.form.get('emergency_phone')
        athlete.status            = request.form.get('status', 'activo')
        athlete.position          = request.form.get('position') or None
        athlete.entry_date        = _parse_date(request.form.get('entry_date')) or date.today()
        athlete.notes             = request.form.get('notes')
        # Sync groups
        AthleteGroup.query.filter_by(athlete_id=athlete.id).delete()
        for gid in request.form.getlist('groups'):
            db.session.add(AthleteGroup(athlete_id=athlete.id, group_id=int(gid)))
        db.session.commit()
        flash('Deportista actualizado.', 'success')
        return redirect(url_for('athletes.detail', id=athlete.id))
    return render_template('athletes/form.html', athlete=athlete, groups=groups)


@athletes_bp.route('/athletes/<int:id>/status', methods=['POST'])
@login_required
@club_required
@permission_required('athletes.manage')
def update_status(id):
    athlete = scoped_or_404(Athlete, id)
    athlete.status = request.form.get('status', athlete.status)
    db.session.commit()
    flash('Estado actualizado.', 'success')
    return redirect(url_for('athletes.detail', id=id))


@athletes_bp.route('/athletes/<int:id>/delete', methods=['POST'])
@login_required
@club_required
@permission_required('athletes.delete')
def delete(id):
    athlete = scoped_or_404(Athlete, id)
    db.session.delete(athlete)
    db.session.commit()
    flash('Deportista eliminado.', 'warning')
    return redirect(url_for('athletes.index'))


@athletes_bp.route('/athletes/<int:id>/upload', methods=['POST'])
@login_required
@club_required
@permission_required('athletes.manage')
def upload_document(id):
    athlete = scoped_or_404(Athlete, id)
    file = request.files.get('file') or request.files.get('document')
    if not file or not file.filename:
        flash('No seleccionaste ningún archivo.', 'danger')
        return redirect(url_for('athletes.detail', id=id))
    if not allowed_file(file.filename):
        flash(f'Tipo de archivo no permitido: {file.filename}. Solo PDF, JPG y PNG.', 'danger')
        return redirect(url_for('athletes.detail', id=id))
    ext = file.filename.rsplit('.', 1)[1].lower()
    from routes.storage import save_file
    saved = save_file(file, f'docs_{athlete.id}', {'pdf', 'jpg', 'jpeg', 'png'},
                      prefix=f'doc_')
    if not saved:
        flash('No se pudo guardar el archivo.', 'danger')
        return redirect(url_for('athletes.detail', id=id))
    doc = AthleteDocument(
        athlete_id=athlete.id, filename=saved,
        original_name=file.filename, file_type=ext,
        description=request.form.get('description', '')
    )
    db.session.add(doc)
    db.session.commit()
    flash('Documento subido correctamente.', 'success')
    return redirect(url_for('athletes.detail', id=id))


@athletes_bp.route('/athletes/<int:id>/export/pdf')
@login_required
@club_required
@permission_required('athletes.view')
def export_pdf(id):
    athlete = scoped_or_404(Athlete, id)
    pdf_bytes = _generate_athlete_pdf(athlete)
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                     download_name=f'deportista_{athlete.id}.pdf')


@athletes_bp.route('/athletes/<int:id>/export/excel')
@login_required
@club_required
@permission_required('athletes.view')
def export_excel(id):
    athlete = scoped_or_404(Athlete, id)
    excel_bytes = _generate_athlete_excel(athlete)
    return send_file(io.BytesIO(excel_bytes),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f'deportista_{athlete.id}.xlsx')




@athletes_bp.route('/athletes/<int:id>/photo', methods=['POST'])
@login_required
@club_required
@permission_required('athletes.manage')
def upload_photo(id):
    """Sube o reemplaza la foto de perfil del deportista."""
    athlete = scoped_or_404(Athlete, id)
    file = request.files.get('photo')
    if not file or not file.filename:
        flash('No seleccionaste ninguna imagen.', 'danger')
        return redirect(url_for('athletes.detail', id=id))
    from routes.storage import save_file, delete_file
    saved = save_file(file, 'photos', {'jpg', 'jpeg', 'png', 'webp'},
                      prefix=f'athlete_{athlete.id}_')
    if not saved:
        flash('La foto debe ser JPG, PNG o WebP.', 'danger')
        return redirect(url_for('athletes.detail', id=id))
    if athlete.photo:
        delete_file(athlete.photo, 'photos')
    athlete.photo = saved
    db.session.commit()
    flash('Foto actualizada.', 'success')
    return redirect(url_for('athletes.detail', id=id))


@athletes_bp.route('/athletes/<int:id>/photo/delete', methods=['POST'])
@login_required
@club_required
@permission_required('athletes.manage')
def delete_photo(id):
    athlete = scoped_or_404(Athlete, id)
    if athlete.photo:
        from routes.storage import delete_file
        delete_file(athlete.photo, 'photos')
        athlete.photo = None
        db.session.commit()
        flash('Foto eliminada.', 'warning')
    return redirect(url_for('athletes.detail', id=id))




# ═════════════════════════════════════════════════════════════════════════════
# IMPORTACIÓN MASIVA (Excel / CSV / PDF)
# ═════════════════════════════════════════════════════════════════════════════

@athletes_bp.route('/athletes/import')
@login_required
@club_required
@permission_required('athletes.manage')
def import_page():
    return render_template('athletes/import.html', preview=None, errors=None, token=None)


@athletes_bp.route('/athletes/import/template')
@login_required
@club_required
@permission_required('athletes.manage')
def import_template():
    from routes.athlete_import import generate_template
    return send_file(io.BytesIO(generate_template()),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='plantilla_deportistas.xlsx', as_attachment=True)


@athletes_bp.route('/athletes/import/preview', methods=['POST'])
@login_required
@club_required
@permission_required('athletes.manage')
def import_preview():
    from routes.athlete_import import parse_file
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Selecciona un archivo.', 'danger')
        return redirect(url_for('athletes.import_page'))

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'csv', 'pdf'):
        flash('Formato no soportado. Usa Excel (.xlsx), CSV o PDF.', 'danger')
        return redirect(url_for('athletes.import_page'))

    data = file.read()
    if len(data) > 10 * 1024 * 1024:
        flash('El archivo supera 10 MB.', 'danger')
        return redirect(url_for('athletes.import_page'))

    rows, errors = parse_file(file.filename, data)

    if not rows and errors:
        for e in errors[:5]:
            flash(e, 'danger')
        return redirect(url_for('athletes.import_page'))

    # Detectar duplicados dentro del club (por número de documento)
    club_docs = {a.document_number for a in
                 Athlete.query.filter_by(club_id=current_club_id()).all()
                 if a.document_number}
    for r in rows:
        r['duplicate'] = bool(r.get('document_number') and
                              r['document_number'] in club_docs)

    # Guardar archivo temporal para el paso de confirmación
    token = uuid.uuid4().hex
    tmp_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'tmp_imports')
    os.makedirs(tmp_dir, exist_ok=True)
    with open(os.path.join(tmp_dir, f'{token}.{ext}'), 'wb') as f:
        f.write(data)

    return render_template('athletes/import.html', preview=rows, errors=errors,
                           token=token, ext=ext,
                           new_count=sum(1 for r in rows if not r['duplicate']),
                           dup_count=sum(1 for r in rows if r['duplicate']))


@athletes_bp.route('/athletes/import/confirm', methods=['POST'])
@login_required
@club_required
@permission_required('athletes.manage')
def import_confirm():
    from routes.athlete_import import parse_file
    from datetime import datetime as _dt

    token = request.form.get('token', '')
    ext = request.form.get('ext', '')
    if not token or not ext or not token.isalnum():
        flash('Sesión de importación inválida. Sube el archivo de nuevo.', 'danger')
        return redirect(url_for('athletes.import_page'))

    tmp_path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'tmp_imports',
                            f'{token}.{ext}')
    if not os.path.exists(tmp_path):
        flash('El archivo temporal expiró. Sube el archivo de nuevo.', 'danger')
        return redirect(url_for('athletes.import_page'))

    with open(tmp_path, 'rb') as f:
        data = f.read()
    rows, _ = parse_file(f'import.{ext}', data)

    club_docs = {a.document_number for a in
                 Athlete.query.filter_by(club_id=current_club_id()).all()
                 if a.document_number}
    skip_duplicates = request.form.get('skip_duplicates') == 'on'

    created, skipped = 0, 0
    for r in rows:
        doc = r.get('document_number') or None
        if doc and doc in club_docs:
            if skip_duplicates:
                skipped += 1
                continue
        birth = None
        if r.get('birth_date'):
            try:
                birth = _dt.strptime(r['birth_date'], '%Y-%m-%d').date()
            except ValueError:
                birth = None
        athlete = Athlete(
            club_id=current_club_id(),
            first_name=r['first_name'], last_name=r['last_name'],
            document_type=r.get('document_type') or 'CC',
            document_number=doc,
            birth_date=birth,
            gender=r.get('gender') or None,
            phone=r.get('phone') or None,
            email=r.get('email') or None,
            address=r.get('address') or None,
            city=r.get('city') or None,
            emergency_contact=r.get('emergency_contact') or None,
            emergency_phone=r.get('emergency_phone') or None,
            status=r.get('status') or 'activo',
            position=r.get('position') or None,
            notes=r.get('notes') or None,
        )
        db.session.add(athlete)
        if doc:
            club_docs.add(doc)
        created += 1
    db.session.commit()

    # Limpiar archivo temporal
    try:
        os.remove(tmp_path)
    except Exception:
        pass

    from models import log_activity
    log_activity('Importación de deportistas',
                 f'{created} creados, {skipped} omitidos', user=current_user)
    flash(f'Importación completada: {created} deportista(s) creado(s)'
          + (f', {skipped} duplicado(s) omitido(s).' if skipped else '.'), 'success')
    return redirect(url_for('athletes.index'))


# ── Helpers ──────────────────────────────────────────────────────────────────
def _parse_date(date_str):
    if not date_str:
        return None
    try:
        from datetime import datetime
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return None


def _generate_athlete_pdf(athlete):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, Image as RLImage)
    from reportlab.lib.units import cm
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Foto de perfil (si existe)
    if athlete.photo:
        from routes.storage import fetch_bytes
        photo_bytes = fetch_bytes(athlete.photo, 'photos')
        if photo_bytes:
            try:
                img = RLImage(io.BytesIO(photo_bytes), width=3.5*cm, height=3.5*cm,
                              kind='proportional')
                img.hAlign = 'CENTER'
                story.append(img)
                story.append(Spacer(1, 0.3*cm))
            except Exception:
                pass

    story.append(Paragraph(f"Ficha Deportista: {athlete.full_name}", styles['Title']))
    story.append(Spacer(1, 0.5*cm))
    data = [
        ['Documento',  f"{athlete.document_type} {athlete.document_number or '-'}"],
        ['Fecha Nac.', str(athlete.birth_date or '-')],
        ['Género',     athlete.gender or '-'],
        ['Posición',   athlete.position or '-'],
        ['Teléfono',   athlete.phone or '-'],
        ['Email',      athlete.email or '-'],
        ['Ciudad',     athlete.city or '-'],
        ['Estado',     athlete.status],
        ['Deuda',      f"${athlete.total_debt:,.0f}"],
    ]
    t = Table(data, colWidths=[5*cm, 10*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#1a56db')),
        ('TEXTCOLOR',  (0,0), (0,-1), colors.white),
        ('FONTNAME',   (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('ROWBACKGROUNDS', (1,0), (-1,-1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def _generate_athlete_excel(athlete):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = 'Deportista'
    header_fill = PatternFill('solid', fgColor='1a56db')
    header_font = Font(color='FFFFFF', bold=True)

    start_row = 1
    # Foto de perfil (si existe)
    if athlete.photo:
        from routes.storage import fetch_bytes
        photo_bytes = fetch_bytes(athlete.photo, 'photos')
        if photo_bytes:
            try:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(io.BytesIO(photo_bytes))
                # Escalar a máx 140px conservando proporción
                ratio = min(140 / img.width, 140 / img.height, 1)
                img.width  = int(img.width * ratio)
                img.height = int(img.height * ratio)
                ws.add_image(img, 'B1')
                # Dejar espacio para la foto
                for _ in range(8):
                    ws.append([])
                start_row = 9
            except Exception:
                pass

    ws.append(['Campo', 'Valor'])
    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = header_font
    rows = [
        ('Nombre',    athlete.full_name),
        ('Documento', f"{athlete.document_type} {athlete.document_number or ''}"),
        ('Fecha Nac.', str(athlete.birth_date or '')),
        ('Género',    athlete.gender or ''),
        ('Posición',  athlete.position or ''),
        ('Teléfono',  athlete.phone or ''),
        ('Email',     athlete.email or ''),
        ('Ciudad',    athlete.city or ''),
        ('Estado',    athlete.status),
        ('Deuda',     athlete.total_debt),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 32
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
