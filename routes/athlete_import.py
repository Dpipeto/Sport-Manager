"""
Importación masiva de deportistas desde Excel (.xlsx), CSV o PDF.

- Excel/CSV: una fila por deportista, con encabezados flexibles
  (reconoce variantes: "Nombres", "nombre", "first_name", etc.)
- PDF: extrae texto y busca campos etiquetados tipo "Nombres: Juan"
  (formato típico de fichas de inscripción). Detecta varios deportistas
  si el PDF tiene varias fichas.
"""
import re
import io
from datetime import datetime, date


# ── Mapeo flexible de encabezados ────────────────────────────────────────────
HEADER_MAP = {
    'first_name': ['nombres', 'nombre', 'first name', 'first_name', 'primer nombre'],
    'last_name':  ['apellidos', 'apellido', 'last name', 'last_name'],
    'document_type':   ['tipo documento', 'tipo doc', 'tipo_documento', 'tipo de documento'],
    'document_number': ['numero documento', 'documento', 'número documento', 'num documento',
                        'numero_documento', 'nro documento', 'identificacion', 'identificación',
                        'cedula', 'cédula', 'ti', 'cc'],
    'birth_date': ['fecha nacimiento', 'nacimiento', 'fecha de nacimiento', 'birth_date',
                   'fecha_nacimiento', 'f. nacimiento'],
    'gender': ['genero', 'género', 'sexo', 'gender'],
    'phone': ['telefono', 'teléfono', 'celular', 'phone', 'tel'],
    'email': ['email', 'correo', 'correo electronico', 'correo electrónico', 'e-mail'],
    'address': ['direccion', 'dirección', 'address'],
    'city': ['ciudad', 'city', 'municipio'],
    'emergency_contact': ['contacto emergencia', 'contacto de emergencia', 'acudiente',
                          'emergencia contacto', 'nombre acudiente'],
    'emergency_phone': ['telefono emergencia', 'teléfono emergencia', 'tel emergencia',
                        'celular acudiente', 'telefono acudiente', 'teléfono acudiente'],
    'position': ['posicion', 'posición', 'position', 'posicion de juego', 'posición de juego', 'puesto'],
    'status': ['estado', 'status'],
    'notes': ['notas', 'observaciones', 'notes', 'comentarios'],
}

TEMPLATE_COLUMNS = [
    ('Nombres *',            'first_name'),
    ('Apellidos *',          'last_name'),
    ('Tipo Documento',       'document_type'),
    ('Número Documento',     'document_number'),
    ('Fecha Nacimiento',     'birth_date'),
    ('Género',               'gender'),
    ('Teléfono',             'phone'),
    ('Email',                'email'),
    ('Dirección',            'address'),
    ('Ciudad',               'city'),
    ('Contacto Emergencia',  'emergency_contact'),
    ('Teléfono Emergencia',  'emergency_phone'),
    ('Posición',             'position'),
    ('Estado',               'status'),
    ('Notas',                'notes'),
]


def _norm(text):
    """Normaliza un encabezado para comparación."""
    if text is None:
        return ''
    return re.sub(r'[^a-z0-9 ]', '', str(text).strip().lower()
                  .replace('á','a').replace('é','e').replace('í','i')
                  .replace('ó','o').replace('ú','u').replace('ñ','n'))


def _match_header(header):
    """Retorna el campo interno correspondiente al encabezado, o None."""
    h = _norm(header).replace('*', '').strip()
    for field, variants in HEADER_MAP.items():
        for v in variants:
            if h == _norm(v):
                return field
    return None


def _parse_date_value(value):
    """Acepta date, datetime, o strings DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY."""
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _clean_row(row):
    """Limpia y valida una fila parseada. Retorna (row, error)."""
    row = {k: (str(v).strip() if v is not None else '') for k, v in row.items()}
    if not row.get('first_name') or not row.get('last_name'):
        return None, 'Faltan nombres o apellidos'
    # Fecha
    if row.get('birth_date'):
        parsed = _parse_date_value(row['birth_date'])
        row['birth_date'] = parsed or ''
    # Estado válido
    st = row.get('status', '').lower()
    row['status'] = st if st in ('activo', 'inactivo', 'suspendido') else 'activo'
    # Tipo doc por defecto
    if not row.get('document_type'):
        row['document_type'] = 'CC'
    return row, None


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL
# ═════════════════════════════════════════════════════════════════════════════

def parse_excel(file_bytes):
    """Parsea un .xlsx. Retorna (rows, errors)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # Encabezados
    headers = None
    for raw in rows_iter:
        if raw and any(c is not None and str(c).strip() for c in raw):
            headers = [_match_header(c) for c in raw]
            break
    if not headers or not any(headers):
        return [], ['No se reconocieron los encabezados. Usa la plantilla descargable.']

    rows, errors = [], []
    for i, raw in enumerate(rows_iter, start=2):
        if not raw or not any(c is not None and str(c).strip() for c in raw):
            continue
        row = {}
        for j, field in enumerate(headers):
            if field and j < len(raw):
                row[field] = raw[j]
        clean, err = _clean_row(row)
        if err:
            errors.append(f'Fila {i}: {err}')
        else:
            rows.append(clean)
    return rows, errors


# ═════════════════════════════════════════════════════════════════════════════
# CSV
# ═════════════════════════════════════════════════════════════════════════════

def parse_csv(file_bytes):
    """Parsea CSV (detecta , o ;). Retorna (rows, errors)."""
    import csv as csv_mod
    text = file_bytes.decode('utf-8-sig', errors='replace')
    delimiter = ';' if text.count(';') > text.count(',') else ','
    reader = csv_mod.reader(io.StringIO(text), delimiter=delimiter)

    headers = None
    rows, errors = [], []
    for i, raw in enumerate(reader, start=1):
        if not raw or not any(c.strip() for c in raw):
            continue
        if headers is None:
            headers = [_match_header(c) for c in raw]
            if not any(headers):
                return [], ['No se reconocieron los encabezados. Usa la plantilla descargable.']
            continue
        row = {}
        for j, field in enumerate(headers):
            if field and j < len(raw):
                row[field] = raw[j]
        clean, err = _clean_row(row)
        if err:
            errors.append(f'Fila {i}: {err}')
        else:
            rows.append(clean)
    return rows, errors


# ═════════════════════════════════════════════════════════════════════════════
# PDF — campos etiquetados tipo ficha de inscripción
# ═════════════════════════════════════════════════════════════════════════════

PDF_FIELD_PATTERNS = {
    'first_name':        r'(?:nombres?)\s*[:：]\s*(.+)',
    'last_name':         r'(?:apellidos?)\s*[:：]\s*(.+)',
    'document_number':   r'(?:n[uú]mero\s+(?:de\s+)?documento|documento|c[eé]dula|identificaci[oó]n|t\.?i\.?|c\.?c\.?)\s*[:：]\s*(\d[\d.\- ]*)',
    'document_type':     r'(?:tipo\s+(?:de\s+)?documento)\s*[:：]\s*(\w+)',
    'birth_date':        r'(?:fecha\s+(?:de\s+)?nacimiento|nacimiento)\s*[:：]\s*([\d/\-]+)',
    'gender':            r'(?:g[eé]nero|sexo)\s*[:：]\s*(\w+)',
    'phone':             r'(?:tel[eé]fono|celular)\s*[:：]\s*([+(]?\d[\d\- +()]*)',
    'email':             r'(?:correo(?:\s+electr[oó]nico)?|email|e-mail)\s*[:：]\s*(\S+@\S+)',
    'address':           r'(?:direcci[oó]n)\s*[:：]\s*(.+)',
    'city':              r'(?:ciudad|municipio)\s*[:：]\s*(.+)',
    'emergency_contact': r'(?:contacto\s+(?:de\s+)?emergencia|acudiente)\s*[:：]\s*(.+)',
    'emergency_phone':   r'(?:tel[eé]fono\s+(?:de\s+)?emergencia|celular\s+acudiente|tel[eé]fono\s+acudiente)\s*[:：]\s*([+(]?\d[\d\- +()]*)',
    'position':          r'(?:posici[oó]n(?:\s+de\s+juego)?|puesto)\s*[:：]\s*(.+)',
}


def parse_pdf(file_bytes):
    """
    Extrae texto del PDF y busca campos etiquetados.
    Si el PDF tiene varias fichas (varios "Nombres:"), separa por bloques.
    Retorna (rows, errors).
    """
    try:
        from pypdf import PdfReader
    except ImportError:
        return [], ['Falta la librería pypdf. Instálala con: pip install pypdf']

    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        text = '\n'.join((page.extract_text() or '') for page in reader.pages)
    except Exception as e:
        return [], [f'No se pudo leer el PDF: {e}']

    if not text.strip():
        return [], ['El PDF no contiene texto extraíble (¿es un escaneo? '
                    'Los PDF escaneados como imagen no son compatibles; usa Excel).']

    # Dividir en bloques: cada "Nombres:" inicia una ficha nueva
    starts = [m.start() for m in re.finditer(r'(?im)^\s*nombres?\s*[:：]', text)]
    if not starts:
        return [], ['No se encontraron campos tipo "Nombres:" en el PDF. '
                    'El PDF debe tener los datos etiquetados (Nombres:, Apellidos:, etc.) '
                    'o usa la plantilla Excel.']

    blocks = []
    for i, s in enumerate(starts):
        end = starts[i + 1] if i + 1 < len(starts) else len(text)
        blocks.append(text[s:end])

    rows, errors = [], []
    for i, block in enumerate(blocks, start=1):
        row = {}
        for field, pattern in PDF_FIELD_PATTERNS.items():
            m = re.search(pattern, block, re.IGNORECASE)
            if m:
                value = m.group(1).strip().split('\n')[0].strip()
                # Limpiar números de documento
                if field == 'document_number':
                    value = re.sub(r'[.\s\-]', '', value)
                row[field] = value
        clean, err = _clean_row(row)
        if err:
            errors.append(f'Ficha {i}: {err}')
        else:
            rows.append(clean)
    return rows, errors


# ═════════════════════════════════════════════════════════════════════════════
# DISPATCHER
# ═════════════════════════════════════════════════════════════════════════════

def parse_file(filename, file_bytes):
    """Detecta el tipo por extensión y parsea. Retorna (rows, errors)."""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext == 'xlsx':
        return parse_excel(file_bytes)
    if ext == 'csv':
        return parse_csv(file_bytes)
    if ext == 'pdf':
        return parse_pdf(file_bytes)
    return [], [f'Formato .{ext} no soportado. Usa .xlsx, .csv o .pdf']


def generate_template():
    """Genera la plantilla Excel descargable con ejemplo. Retorna bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Deportistas'

    headers = [label for label, _ in TEMPLATE_COLUMNS]
    ws.append(headers)
    fill = PatternFill('solid', fgColor='1a56db')
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    # Fila de ejemplo
    ws.append(['Juan Camilo', 'Pérez Gómez', 'TI', '1023456789', '15/03/2012',
               'Masculino', '3001234567', 'juan@correo.com', 'Calle 10 # 5-20',
               'Chía', 'María Gómez', '3109876543', 'Delantero', 'activo', 'Alergia al maní'])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
