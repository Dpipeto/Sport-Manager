from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_file
from flask_login import login_required, current_user
from models import Payment, Athlete, Group
from app import db, mail
from datetime import date, datetime
from sqlalchemy import func
from routes.decorators import permission_required, club_required, current_club_id, athlete_scoped_or_404
import io

billing_bp = Blueprint('billing', __name__)


# ── Lista (admin + recepcion) ─────────────────────────────────────────────────
@billing_bp.route('/billing')
@login_required
@club_required
@permission_required('billing.view')
def index():
    search = request.args.get('search', '')
    status = request.args.get('status', '')
    ptype  = request.args.get('type', '')
    query  = Payment.query.join(Athlete).filter(Athlete.club_id == current_club_id())
    if search:
        query = query.filter(
            (Athlete.first_name.ilike(f'%{search}%')) |
            (Athlete.last_name.ilike(f'%{search}%'))
        )
    if status:
        query = query.filter(Payment.status == status)
    if ptype:
        query = query.filter(Payment.type == ptype)
    payments = query.order_by(Payment.date.desc()).all()
    groups = Group.query.filter_by(active=True, club_id=current_club_id()).all()
    # KPIs
    club_filter = Payment.query.join(Athlete).filter(Athlete.club_id == current_club_id())
    total_charged = sum(p.amount for p in club_filter.filter(Payment.type == 'cargo').all())
    total_paid    = sum(p.amount for p in club_filter.filter(Payment.type == 'pago').all())
    total_pending = sum(p.amount for p in club_filter.filter(
        Payment.type == 'cargo', Payment.status == 'pendiente').all())
    return render_template('billing/index.html', payments=payments, groups=groups,
                           search=search, status_filter=status, type_filter=ptype,
                           total_charged=total_charged, total_paid=total_paid,
                           total_pending=total_pending)


@billing_bp.route('/billing/new', methods=['GET', 'POST'])
@login_required
@club_required
@permission_required('billing.view')
def new():
    athletes = Athlete.query.filter_by(club_id=current_club_id()).order_by(Athlete.last_name).all()
    if request.method == 'POST':
        send_email_flag = request.form.get('send_email') == 'on'
        payment = Payment(
            athlete_id=int(request.form.get('athlete_id')),
            type=request.form.get('type', 'cargo'),
            concept=request.form.get('concept'),
            amount=float(request.form.get('amount', 0)),
            date=_parse_date(request.form.get('date')) or date.today(),
            due_date=_parse_date(request.form.get('due_date')),
            status=request.form.get('status', 'pendiente'),
            payment_method=request.form.get('payment_method'),
            reference=request.form.get('reference'),
            notes=request.form.get('notes'),
        )
        db.session.add(payment)
        db.session.commit()
        if send_email_flag:
            _send_payment_email(payment)
            payment.email_sent = True
            db.session.commit()
        flash('Registro creado exitosamente.', 'success')
        return redirect(url_for('billing.index'))
    return render_template('billing/form.html', payment=None, athletes=athletes)


@billing_bp.route('/billing/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@club_required
@permission_required('billing.view')
def edit(id):
    payment  = athlete_scoped_or_404(Payment, id)
    athletes = Athlete.query.filter_by(club_id=current_club_id()).order_by(Athlete.last_name).all()
    if request.method == 'POST':
        payment.athlete_id     = int(request.form.get('athlete_id'))
        payment.type           = request.form.get('type', 'cargo')
        payment.concept        = request.form.get('concept')
        payment.amount         = float(request.form.get('amount', 0))
        payment.date           = _parse_date(request.form.get('date')) or date.today()
        payment.due_date       = _parse_date(request.form.get('due_date'))
        payment.status         = request.form.get('status', 'pendiente')
        payment.payment_method = request.form.get('payment_method')
        payment.reference      = request.form.get('reference')
        payment.notes          = request.form.get('notes')
        db.session.commit()
        flash('Registro actualizado.', 'success')
        return redirect(url_for('billing.index'))
    return render_template('billing/form.html', payment=payment, athletes=athletes)


@billing_bp.route('/billing/<int:id>/delete', methods=['POST'])
@login_required
@club_required
@permission_required('billing.view')
def delete(id):
    payment = athlete_scoped_or_404(Payment, id)
    db.session.delete(payment)
    db.session.commit()
    flash('Registro eliminado.', 'warning')
    return redirect(url_for('billing.index'))


@billing_bp.route('/billing/<int:id>/send-email', methods=['POST'])
@login_required
@club_required
@permission_required('billing.view')
def send_email(id):
    payment = athlete_scoped_or_404(Payment, id)
    if _send_payment_email(payment):
        payment.email_sent = True
        db.session.commit()
        flash('Email enviado correctamente.', 'success')
    else:
        flash('No se pudo enviar el email. Verifica la configuración de correo.', 'danger')
    return redirect(url_for('billing.index'))


@billing_bp.route('/billing/bulk-charge', methods=['POST'])
@login_required
@club_required
@permission_required('billing.view')
def bulk_charge():
    group_id   = request.form.get('group_id')
    concept    = request.form.get('concept', 'Mensualidad')
    amount     = float(request.form.get('amount', 0))
    due_date   = _parse_date(request.form.get('due_date'))
    send_emails = request.form.get('send_emails') == 'on'
    if not group_id:
        flash('Selecciona un grupo.', 'danger')
        return redirect(url_for('billing.index'))
    group = Group.query.get_or_404(int(group_id))
    if group.club_id != current_club_id():
        flash('Grupo no válido.', 'danger')
        return redirect(url_for('billing.index'))
    count = 0
    for ag in group.athletes:
        if ag.active:
            p = Payment(
                athlete_id=ag.athlete_id,
                type='cargo', concept=concept,
                amount=amount, date=date.today(),
                due_date=due_date, status='pendiente',
            )
            db.session.add(p)
            db.session.flush()
            if send_emails:
                _send_payment_email(p)
                p.email_sent = True
            count += 1
    db.session.commit()
    flash(f'Cobro masivo generado para {count} deportistas del grupo {group.name}.', 'success')
    return redirect(url_for('billing.index'))


@billing_bp.route('/billing/report/pdf')
@login_required
@club_required
@permission_required('billing.view')
def report_pdf():
    payments = Payment.query.join(Athlete).filter(Athlete.club_id == current_club_id()).order_by(Payment.date.desc()).all()
    pdf_bytes = _generate_billing_pdf(payments)
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                     download_name='reporte_cobros.pdf')


@billing_bp.route('/billing/report/excel')
@login_required
@club_required
@permission_required('billing.view')
def report_excel():
    payments = Payment.query.join(Athlete).filter(Athlete.club_id == current_club_id()).order_by(Payment.date.desc()).all()
    excel_bytes = _generate_billing_excel(payments)
    return send_file(io.BytesIO(excel_bytes),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='reporte_cobros.xlsx')


# ── Helpers ───────────────────────────────────────────────────────────────────
def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


def _send_payment_email(payment):
    try:
        from flask_mail import Message
        athlete = payment.athlete
        if not athlete.email:
            return False
        msg = Message(
            subject=f"Notificación de {'cobro' if payment.type=='cargo' else 'pago'}: {payment.concept}",
            recipients=[athlete.email],
            body=f"Estimado/a {athlete.full_name},\n\n"
                 f"{'Se ha generado un cobro' if payment.type=='cargo' else 'Se ha registrado un pago'} "
                 f"por ${payment.amount:,.0f} - {payment.concept}.\n\n"
                 f"Estado: {payment.status}\n"
                 f"Fecha: {payment.date}\n\n"
                 f"Saludos,\nSportManager"
        )
        mail.send(msg)
        return True
    except Exception:
        return False


def _generate_billing_pdf(payments):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = [Paragraph("Reporte de Cobros", styles['Title']), Spacer(1, 0.5*cm)]
    headers = ['Deportista','Concepto','Tipo','Monto','Fecha','Estado']
    data = [headers] + [
        [p.athlete.full_name, p.concept, p.type,
         f"${p.amount:,.0f}", str(p.date), p.status]
        for p in payments
    ]
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a56db')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def _generate_billing_excel(payments):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cobros'
    header_fill = PatternFill('solid', fgColor='1a56db')
    header_font = Font(color='FFFFFF', bold=True)
    headers = ['Deportista','Concepto','Tipo','Monto','Fecha','Estado','Método','Referencia']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for p in payments:
        ws.append([p.athlete.full_name, p.concept, p.type,
                   p.amount, str(p.date), p.status,
                   p.payment_method or '', p.reference or ''])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
